import os
import re
import json
import queue
import threading
import traceback
import subprocess
import shutil
from urllib.parse import urlparse, unquote
from datetime import datetime
from ftplib import FTP

import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


AWS_PREFIX = "https://ota-dl.vidaahub.com"
FTP_USER = "ftpuser"
FTP_PASS = "Ftp@User2022"

CONFIG_FILE = os.path.join(
    os.path.expanduser("~"),
    ".ota_deploy_tool_config.json"
)

OUTPUT_HEADERS = [
    "设备类型代码",
    "设备类型名称",
    "特征码",
    "内部机型信息",
    "设备扩展信息",
    "OTA类型",
    "源版本",
    "目标完整版本",
    "目标版本",
    "品牌组",
    "区域组",
    "MAC组",
    "定向组",
    "升级文件地址",
    "升级文件的MD5值",
    "文件大小",
    "SHA256",
    "EULA文件地址",
]

COLUMN_ALIASES = {
    "机型信息": ["机型", "机型信息", "内部机型", "model", "model name"],
    "机器扩展码": [
        "机器扩展码", "扩展码", "扩展信息", "机器扩展信息", "设备扩展信息",
        "extension info", "extension code", "feature code",
    ],
    "OTA类型": ["OTA类型", "OTA升级类型", "升级类型", "ota type", "upgrade type"],
    "源版本": ["源版本", "source version", "源文件版本"],
    "目标版本": ["目标版本", "target version", "完整版本", "目标文件版本"],
    "版本变化": ["版本变化", "version change", "版本差异"],
    "升级包地址": ["升级包地址", "整包地址", "整包文件地址", "完整包地址", "全包地址", "OTA整包地址", "ZIP地址", "ZIP文件地址", "升级文件地址", "ftp path", "ftp url", "upgrade url", "file path", "文件路径"],
    "升级文件的MD5值": ["升级文件的md5值", "升级文件md5值", "md5", "md5值", "md5 checksum"],
    "文件大小": ["升级包大小", "升级包大小(byte)", "升级包大小（byte）", "文件大小", "升级文件包大小（byte）", "升级文件包大小(byte)", "file size", "size"],
    "语言信息": ["语言信息", "language", "lang", "语言"],
    "品牌": ["品牌", "brand"],
    "区域": ["区域", "区域组", "地区", "地区组", "region", "area", "country", "market", "territory"],
    "SHA256": ["sha256", "sha256值"],
    "EULA文件地址": ["eula文件地址", "eula", "eula url"],
    "MAC": ["mac", "mac地址", "mac address"],
}


class OTATemplateApp:

    def __init__(self, root):
        self.root = root
        self.root.title("OTA部署文件生成工具")
        self.root.geometry("1180x860")

        self.config = self.load_config()

        self.source_file_var = tk.StringVar(value=self.config.get("source_file", ""))
        self.device_code_var = tk.StringVar(value=self.config.get("device_code", ""))
        self.device_name_var = tk.StringVar(value=self.config.get("device_name", ""))
        self.feature_code_var = tk.StringVar(value=self.config.get("feature_code", ""))
        self.aws_suffix_var = tk.StringVar(value=self.config.get("aws_suffix", ""))
        self.aws_full_location_var = tk.StringVar(value="")
        self.mac_group_var = tk.StringVar(value=self.config.get("mac_group", ""))

        self.device_code_var.trace_add("write", self.update_aws_suffix_from_device_code)
        self.update_aws_suffix_from_device_code()

        self.status_var = tk.StringVar(value="准备就绪")
        self.download_progress_var = tk.DoubleVar(value=0)
        self.download_status_var = tk.StringVar(value="下载进度：0%")

        self.animation_running = False
        self.animation_dots = 0
        self.animation_text = "正在处理"

        self.log_queue = queue.Queue()
        self.source_wb = None
        self.data_sheet_name = None

        self.build_ui()
        self.root.after(100, self.process_log_queue)

    # ==========================================================
    # Config
    # ==========================================================

    def load_config(self):
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def get_default_aws_suffix(self, device_code=None):
        if device_code is None:
            device_code = self.device_code_var.get().strip()
        today = datetime.now().strftime("%Y%m%d")
        return f"/ota/{device_code.strip()}/{today}"

    def get_default_aws_full_location(self, device_code=None):
        """Build the full OTA download location shown in the UI.
        This is informational only and is not editable by the user.
        """
        suffix = self.get_default_aws_suffix(device_code)
        return self.sanitize_url(AWS_PREFIX.rstrip("/") + suffix)

    def update_aws_suffix_from_device_code(self, *args):
        suffix = self.get_default_aws_suffix()
        self.aws_suffix_var.set(suffix)
        self.aws_full_location_var.set(AWS_PREFIX.rstrip("/") + suffix)

    def save_config(self):
        self.update_aws_suffix_from_device_code()
        source_file = self.source_file_var.get().strip()
        data = {
            "source_file": source_file,
            "last_source_dir": os.path.dirname(source_file) if source_file else "",
            "device_code": self.device_code_var.get().strip(),
            "device_name": self.device_name_var.get().strip(),
            "feature_code": self.feature_code_var.get().strip(),
            "aws_suffix": self.aws_suffix_var.get().strip(),
            "mac_group": self.mac_group_var.get().strip(),
        }
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.log_message(f"配置保存失败：{e}")

    # ==========================================================
    # UI
    # ==========================================================

    def build_ui(self):
        frame = tk.Frame(self.root, padx=12, pady=12)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="Source Excel 文件：").grid(row=0, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.source_file_var, width=90).grid(row=0, column=1, padx=6, sticky="we")
        tk.Button(frame, text="选择文件", command=self.select_source_file).grid(row=0, column=2, padx=6)

        tk.Label(frame, text="设备类型代码：").grid(row=1, column=0, sticky="w", pady=6)
        tk.Entry(frame, textvariable=self.device_code_var, width=45).grid(row=1, column=1, sticky="w", padx=6)

        tk.Label(frame, text="设备类型名称：").grid(row=2, column=0, sticky="w", pady=6)
        tk.Entry(frame, textvariable=self.device_name_var, width=45).grid(row=2, column=1, sticky="w", padx=6)

        tk.Label(frame, text="Feature Code / 特征码：").grid(row=3, column=0, sticky="w", pady=6)
        tk.Entry(frame, textvariable=self.feature_code_var, width=45).grid(row=3, column=1, sticky="w", padx=6)

        tk.Label(frame, text="MAC 地址组：").grid(row=4, column=0, sticky="w", pady=6)
        tk.Entry(frame, textvariable=self.mac_group_var, width=70).grid(row=4, column=1, sticky="w", padx=6)

        tk.Label(frame, text="AWS 文件路径：").grid(row=5, column=0, sticky="w", pady=6)

        aws_display_label = tk.Label(
            frame,
            textvariable=self.aws_full_location_var,
            fg="blue",
            bg="#F3F6FA",
            relief=tk.SUNKEN,
            anchor="w",
            padx=8,
            pady=4,
            width=82
        )
        aws_display_label.grid(row=5, column=1, sticky="we", padx=6)

        button_frame = tk.Frame(frame)
        button_frame.grid(row=6, column=0, columnspan=3, pady=18)
        inner_button_frame = tk.Frame(button_frame)
        inner_button_frame.pack()

        self.download_zip_button = tk.Button(
            inner_button_frame, text="下载ZIP文件", command=self.download_zip_only,
            bg="#A4C2F4", fg="black", activebackground="#6FA8DC",
            font=("Arial", 14, "bold"), width=18, height=2, relief=tk.RAISED, bd=4, cursor="hand2"
        )
        self.download_zip_button.grid(row=0, column=0, padx=12)

        self.upload_s3_button = tk.Button(
            inner_button_frame, text="上传ZIP文件到AWS", command=self.upload_to_s3_only,
            bg="#C9A0DC", fg="black", activebackground="#9B59B6",
            font=("Arial", 14, "bold"), width=18, height=2, relief=tk.RAISED, bd=4, cursor="hand2"
        )
        self.upload_s3_button.grid(row=0, column=1, padx=12)

        self.generate_template_button = tk.Button(
            inner_button_frame, text="生成模版文件", command=self.generate_template_only,
            bg="#FFD966", fg="black", activebackground="#F4B183",
            font=("Arial", 14, "bold"), width=18, height=2, relief=tk.RAISED, bd=4, cursor="hand2"
        )
        self.generate_template_button.grid(row=0, column=2, padx=12)

        self.auto_execute_button = tk.Button(
            inner_button_frame, text="自动执行", command=self.auto_execute_all,
            bg="#93C47D", fg="black", activebackground="#6AA84F",
            font=("Arial", 14, "bold"), width=18, height=2, relief=tk.RAISED, bd=4, cursor="hand2"
        )
        self.auto_execute_button.grid(row=0, column=3, padx=12)

        tk.Label(button_frame, textvariable=self.status_var, fg="#555555", font=("Arial", 11)).pack(pady=4)

        progress_frame = tk.Frame(frame)
        progress_frame.grid(row=7, column=0, columnspan=3, sticky="we", pady=8)
        tk.Label(progress_frame, textvariable=self.download_status_var, font=("Arial", 11), fg="#333333").pack()
        self.download_progress_bar = ttk.Progressbar(
            progress_frame, variable=self.download_progress_var, maximum=100, length=700, mode="determinate"
        )
        self.download_progress_bar.pack(pady=4)

        tk.Label(frame, text="执行日志：").grid(row=8, column=0, sticky="nw", pady=6)
        self.log_text = scrolledtext.ScrolledText(frame, height=30, width=145)
        self.log_text.grid(row=8, column=1, columnspan=2, sticky="nsew", padx=6)

        copy_log_button = tk.Button(
            frame,
            text="复制日志 / Copy Log",
            command=self.copy_log_to_clipboard,
            bg="#D9EAD3",
            fg="black",
            font=("Arial", 10, "bold"),
            cursor="hand2"
        )
        copy_log_button.grid(row=9, column=1, sticky="e", padx=6, pady=6)

        self.log_text.bind("<Command-a>", self.select_all_log_text)
        self.log_text.bind("<Control-a>", self.select_all_log_text)

        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(8, weight=1)

    # ==========================================================
    # Center Popup
    # ==========================================================

    def show_center_message(self, title, message):
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.transient(self.root)
        popup.grab_set()
        popup.resizable(False, False)

        width = 460
        height = 200
        self.root.update_idletasks()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        x = root_x + (root_width // 2) - (width // 2)
        y = root_y + (root_height // 2) - (height // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(popup, padx=22, pady=22)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text=title, font=("Arial", 16, "bold"), fg="#2F5597").pack(pady=(0, 14))
        tk.Label(frame, text=message, font=("Arial", 12), justify="center", wraplength=400).pack(pady=6)
        tk.Button(
            frame, text="确定", width=12, bg="#93C47D", fg="black",
            font=("Arial", 11, "bold"), command=popup.destroy, cursor="hand2"
        ).pack(pady=16)
        popup.focus_force()

    def show_center_error(self, title, message):
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.transient(self.root)
        popup.grab_set()
        popup.resizable(False, False)

        width = 520
        height = 260
        self.root.update_idletasks()
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        x = root_x + (root_width // 2) - (width // 2)
        y = root_y + (root_height // 2) - (height // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(popup, padx=20, pady=20)
        frame.pack(fill="both", expand=True)
        tk.Label(frame, text=title, font=("Arial", 16, "bold"), fg="red").pack(pady=(0, 12))
        text_box = tk.Text(frame, height=7, wrap="word", font=("Consolas", 10))
        text_box.pack(fill="both", expand=True)
        text_box.insert("1.0", message)
        text_box.config(state="disabled")
        tk.Button(
            frame, text="关闭", width=12, bg="#E06666", fg="black",
            font=("Arial", 11, "bold"), command=popup.destroy, cursor="hand2"
        ).pack(pady=12)
        popup.focus_force()

    # ==========================================================
    # Button / animation / log
    # ==========================================================

    def disable_all_buttons(self):
        self.generate_template_button.config(state=tk.DISABLED)
        self.download_zip_button.config(state=tk.DISABLED)
        self.auto_execute_button.config(state=tk.DISABLED)
        self.upload_s3_button.config(state=tk.DISABLED)

    def enable_all_buttons(self):
        self.generate_template_button.config(state=tk.NORMAL)
        self.download_zip_button.config(state=tk.NORMAL)
        self.auto_execute_button.config(state=tk.NORMAL)
        self.upload_s3_button.config(state=tk.NORMAL)

    def start_animation(self, text="正在处理"):
        self.animation_running = True
        self.animation_dots = 0
        self.animation_text = text
        self.animate_status()

    def stop_animation(self, final_text="完成"):
        self.animation_running = False
        self.status_var.set(final_text)
        self.enable_all_buttons()

    def animate_status(self):
        if not self.animation_running:
            return
        dots = "." * (self.animation_dots % 4)
        self.status_var.set(f"{self.animation_text}{dots}")
        self.animation_dots += 1
        self.root.after(400, self.animate_status)

    def log_message(self, message):
        """Write a detailed timestamped message into the UI execution log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        text = str(message)
        # Keep section dividers visually clean.
        if not text or set(text.strip()) <= {"="}:
            line = text
        else:
            line = f"[{timestamp}] {text}"
        self.log_text.insert(tk.END, line + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def copy_log_to_clipboard(self):
        """Copy all content in the execution log window to the system clipboard."""
        content = self.log_text.get("1.0", tk.END).strip()
        if not content:
            self.show_center_message("提示", "执行日志为空，没有可复制内容。")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        self.root.update()
        self.show_center_message("复制成功", "执行日志已复制到剪贴板。")

    def select_all_log_text(self, event=None):
        self.log_text.tag_add("sel", "1.0", tk.END)
        self.log_text.mark_set("insert", "1.0")
        self.log_text.see("insert")
        return "break"

    def thread_log(self, message):
        self.log_queue.put(("log", message))

    def thread_progress(self, percent, text=None):
        self.log_queue.put(("progress", percent, text))

    def process_log_queue(self):
        try:
            while True:
                item = self.log_queue.get_nowait()
                if item[0] == "log":
                    self.log_message(item[1])
                elif item[0] == "progress":
                    percent = max(0, min(100, float(item[1])))
                    text = item[2]
                    self.download_progress_var.set(percent)
                    self.download_status_var.set(text if text else f"下载进度：{percent:.1f}%")
                elif item[0] == "done":
                    success_count = item[1]
                    self.stop_animation("ZIP 下载完成")
                    self.download_progress_var.set(100)
                    self.download_status_var.set("全部 ZIP 下载任务完成")
                    self.show_center_message("下载完成", f"ZIP 文件下载完成。\n\n成功下载：{success_count} 个文件")
                elif item[0] == "s3_done":
                    success_count = item[1]
                    fail_count = item[2]
                    if fail_count == 0:
                        self.stop_animation("S3 上传完成")
                        self.show_center_message("上传完成", f"所有文件已上传至 S3。\n\n成功：{success_count} 个文件")
                    else:
                        self.stop_animation("S3 上传部分失败")
                        self.show_center_error("上传结果", f"S3 上传完成（有部分失败）。\n\n成功：{success_count} 个\n失败：{fail_count} 个\n\n请查看执行日志了解详情。")
                elif item[0] == "error":
                    error_message = item[1]
                    self.stop_animation("下载失败")
                    self.show_center_error("错误", error_message)
        except queue.Empty:
            pass
        self.root.after(100, self.process_log_queue)

    # ==========================================================
    # File select
    # ==========================================================

    def select_source_file(self):
        initial_dir = self.config.get("last_source_dir", "")
        if not initial_dir or not os.path.exists(initial_dir):
            initial_dir = os.path.expanduser("~")
        file_path = filedialog.askopenfilename(
            title="选择 Source Excel 文件",
            initialdir=initial_dir,
            filetypes=[("Excel Files", "*.xlsx *.xlsm"), ("All Files", "*.*")]
        )
        if file_path:
            self.source_file_var.set(file_path)
            self.config["last_source_dir"] = os.path.dirname(file_path)
            self.save_config()
            self.log_message(f"已选择 Source Excel 文件：{file_path}")

    # ==========================================================
    # Column detection
    # ==========================================================

    def normalize_header(self, value):
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = text.replace(" ", "").replace("_", "").replace("-", "")
        text = text.replace("（", "(").replace("）", ")")
        return text

    def detect_columns_by_header(self, ws):
        header_row = None
        best_score = -1
        best_mapping = {}
        for row_idx in range(1, min(ws.max_row, 20) + 1):
            raw_headers = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            normalized_headers = [self.normalize_header(h) for h in raw_headers]
            mapping = {}
            for target_field, aliases in COLUMN_ALIASES.items():
                for alias in aliases:
                    alias_norm = self.normalize_header(alias)
                    for col_idx, header_norm in enumerate(normalized_headers, start=1):
                        if not header_norm:
                            continue
                        if header_norm == alias_norm or alias_norm in header_norm or header_norm in alias_norm:
                            mapping[target_field] = col_idx
                            break
                    if target_field in mapping:
                        break
            if len(mapping) > best_score:
                best_score = len(mapping)
                header_row = row_idx
                best_mapping = mapping
        return header_row, best_mapping, best_score

    def looks_like_version(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^\d{4}[A-Z]{1,2}\d{4}$", text, re.IGNORECASE))

    def looks_like_md5(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^[a-fA-F0-9]{32}$", text))

    def looks_like_size(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^\d{6,}$", text))

    def looks_like_url_zip(self, value):
        text = str(value or "").strip().lower()
        return "://" in text and ".zip" in text

    def looks_like_model(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^\d{2}[A-Z0-9]+(?:\(.+\))?$", text))

    def looks_like_feature_code(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^\d{18,}$", text))

    def looks_like_extension(self, value):
        text = str(value or "").strip()
        return bool(re.match(r"^\d{2}[A-Z0-9]+[A-Z0-9]{4,}$", text)) and not self.looks_like_model(text)

    def detect_columns_without_header(self, ws):
        scores = {}
        for col_idx in range(1, ws.max_column + 1):
            col_scores = {
                "机型信息": 0, "机器扩展码": 0, "OTA类型": 0, "源版本": 0, "目标版本": 0,
                "版本变化": 0, "升级包地址": 0, "升级文件的MD5值": 0, "文件大小": 0, "品牌": 0, "区域": 0, "MAC": 0,
            }
            for row_idx in range(1, min(ws.max_row, 30) + 1):
                text = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
                if not text:
                    continue
                if self.looks_like_model(text):
                    col_scores["机型信息"] += 3
                if self.looks_like_feature_code(text):
                    col_scores["机器扩展码"] += 1
                if self.looks_like_extension(text):
                    col_scores["机器扩展码"] += 3
                if text in ["正常升级", "强制升级", "FTE强制升级"]:
                    col_scores["OTA类型"] += 5
                if self.looks_like_version(text):
                    col_scores["源版本"] += 2
                    col_scores["目标版本"] += 2
                if "_to_" in text:
                    col_scores["版本变化"] += 5
                if self.looks_like_url_zip(text):
                    col_scores["升级包地址"] += 6
                if self.looks_like_md5(text):
                    col_scores["升级文件的MD5值"] += 6
                if self.looks_like_size(text):
                    col_scores["文件大小"] += 4
                if text.lower() in ["hisense", "toshiba", "vidaa"]:
                    col_scores["品牌"] += 5
                if text.upper() in ["US", "USA", "EU", "UK", "CA", "AU", "APAC", "LATAM", "MENA", "GLOBAL"] or text.lower() in ["north america", "europe", "global"]:
                    col_scores["区域"] += 4
                if ("、" in text or "," in text) and re.search(r"[a-fA-F0-9]{8}", text):
                    col_scores["MAC"] += 3
            scores[col_idx] = col_scores

        mapping = {}
        used_cols = set()
        field_priority = ["升级包地址", "升级文件的MD5值", "文件大小", "OTA类型", "版本变化", "品牌", "区域", "MAC", "机型信息", "机器扩展码"]
        for field in field_priority:
            best_col = None
            best_score = 0
            for col_idx, col_scores in scores.items():
                if col_idx in used_cols:
                    continue
                if col_scores.get(field, 0) > best_score:
                    best_score = col_scores[field]
                    best_col = col_idx
            if best_col and best_score > 0:
                mapping[field] = best_col
                used_cols.add(best_col)

        version_cols = []
        for col_idx, col_scores in scores.items():
            if col_idx in used_cols:
                continue
            version_score = max(col_scores.get("源版本", 0), col_scores.get("目标版本", 0))
            if version_score > 0:
                version_cols.append(col_idx)
        version_cols = sorted(version_cols)
        if len(version_cols) >= 1:
            mapping["源版本"] = version_cols[0]
        if len(version_cols) >= 2:
            mapping["目标版本"] = version_cols[1]
        return 0, mapping, len(mapping)

    # ==========================================================
    # Excel read
    # ==========================================================

    def read_source_excel(self, file_path, use_gui_log=True):
        if use_gui_log:
            self.log_message("正在读取 Excel...")
            self.log_message(f"Source Excel 路径：{file_path}")
            self.log_message(f"文件大小：{os.path.getsize(file_path)} bytes")
        self.source_wb = load_workbook(file_path, data_only=False)
        if use_gui_log:
            self.log_message("Excel 工作簿打开成功。")
            self.log_message(f"工作表数量：{len(self.source_wb.worksheets)}")
            self.log_message("工作表列表：" + ", ".join([ws.title for ws in self.source_wb.worksheets]))

        best_ws = None
        best_header_row = None
        best_mapping = {}
        best_score = -1
        best_has_header = True

        for ws in self.source_wb.worksheets:
            header_row, mapping, score = self.detect_columns_by_header(ws)
            no_header_row, no_header_mapping, no_header_score = self.detect_columns_without_header(ws)
            if use_gui_log:
                self.log_message(f"检测工作表：{ws.title}")
                self.log_message(f"  尺寸：{ws.max_row} 行 x {ws.max_column} 列")
                self.log_message(f"  表头识别字段数量：{score}")
                self.log_message(f"  表头模式字段映射：{self.format_mapping_for_log(mapping)}")
                self.log_message(f"  无表头识别字段数量：{no_header_score}")
                self.log_message(f"  无表头模式字段映射：{self.format_mapping_for_log(no_header_mapping)}")
            if no_header_score > score:
                current_score = no_header_score
                current_mapping = no_header_mapping
                current_header_row = no_header_row
                current_has_header = False
            else:
                current_score = score
                current_mapping = mapping
                current_header_row = header_row
                current_has_header = True
            if current_score > best_score:
                best_ws = ws
                best_header_row = current_header_row
                best_mapping = current_mapping
                best_score = current_score
                best_has_header = current_has_header

        if best_ws is None or best_score <= 0:
            raise ValueError("未识别到有效工作表。")

        self.data_sheet_name = best_ws.title
        start_row = best_header_row + 1 if best_has_header else 1

        if use_gui_log:
            self.log_message(f"最终选择数据工作表：{best_ws.title}")
            self.log_message(f"识别模式：{'有表头模式' if best_has_header else '无表头自动识别模式'}")
            if best_has_header:
                self.log_message(f"表头所在行：{best_header_row}")
            else:
                self.log_message("源文件未识别到表头：数据将从第 1 行开始读取。")
            self.log_message("字段映射结果：")
            for field, col_idx in best_mapping.items():
                self.log_message(f"  {field} -> {get_column_letter(col_idx)}列 / 示例值：{best_ws.cell(start_row, col_idx).value}")
            self.log_required_field_status(best_mapping)

        data_rows = []
        for row_idx in range(start_row, best_ws.max_row + 1):
            row_data = {}
            has_any_value = False

            # Keep a copy of the whole original Excel row.
            # This is important because some customer source files use changing
            # header names for the FTP ZIP URL column, for example 整包地址 / 整包文件 / 升级包地址.
            # Even if the normal field mapping misses that column, generation can still
            # scan this row and extract the exact .zip filename.
            all_cell_values = []
            for col_idx in range(1, best_ws.max_column + 1):
                raw_value = best_ws.cell(row=row_idx, column=col_idx).value
                if raw_value not in (None, ""):
                    has_any_value = True
                    all_cell_values.append(str(raw_value).strip())

            for field, col_idx in best_mapping.items():
                value = best_ws.cell(row=row_idx, column=col_idx).value
                row_data[field] = "" if value is None else str(value).strip()

            if has_any_value:
                row_data["_row_number"] = row_idx
                row_data["_all_cell_values"] = all_cell_values
                data_rows.append(row_data)

        if use_gui_log:
            self.log_message(f"读取到的数据行数：{len(data_rows)}")
            self.log_message("Excel 读取完成。")
        return data_rows, best_mapping

    # ==========================================================
    # FTP
    # ==========================================================

    def get_unique_ftp_urls(self, rows):
        """Return unique FTP ZIP URLs from parsed source rows.

        This keeps compatibility with the normal column mapping path.
        The source Excel may call this field 整包地址 / 升级包地址 / 升级文件地址.
        """
        ftp_urls = []
        for row in rows:
            url = self.get_value(row, "升级包地址")
            if url and str(url).strip().lower().startswith("ftp://") and ".zip" in str(url).lower():
                ftp_urls.append(str(url).strip())
        return list(dict.fromkeys(ftp_urls))

    def scan_ftp_zip_urls_from_excel(self, source_file_path):
        """Scan the whole workbook for FTP .zip URLs.

        This is intentionally independent from header detection.  Some source
        files use different column names such as 整包地址, and the safest
        behavior for the download step is to scan every cell and extract any
        ftp://...zip value.
        """
        self.thread_log("开始扫描 Source Excel 中所有单元格，查找 FTP ZIP 地址...")
        self.thread_log(f"扫描文件：{source_file_path}")
        wb = load_workbook(source_file_path, data_only=True, read_only=True)
        urls = []
        for ws in wb.worksheets:
            self.thread_log(f"扫描工作表：{ws.title}，尺寸：{ws.max_row} 行 x {ws.max_column} 列")
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if text.lower().startswith("ftp://") and ".zip" in text.lower():
                        urls.append(text)
                        self.thread_log(f"发现 FTP ZIP 地址：工作表={ws.title}, 单元格={cell.coordinate}")
                        self.thread_log(f"  原始地址：{text}")
        wb.close()
        unique_urls = list(dict.fromkeys(urls))
        self.thread_log(f"扫描完成：发现 FTP ZIP 地址 {len(urls)} 条，去重后 {len(unique_urls)} 个文件。")
        return unique_urls

    def parse_ftp_url(self, ftp_url):
        """Parse FTP URL from source Excel.

        Excel usually stores URLs such as:
        ftp://ftpuser@10.18.203.204/upload/.../32S60S%289A%29_User/.../USBOTA_xxx.zip

        Important: urlparse keeps the path percent-encoded.  FTP servers store
        the real directory name with characters such as parentheses, not with
        literal %28/%29.  Therefore we must unquote the path before cwd/RETR.
        Otherwise ftp.cwd('/upload/.../32S60S%289A%29_User/...') returns:
        ftplib.error_perm: 550 Failed to change directory.
        """
        raw_url = str(ftp_url).strip()
        parsed = urlparse(raw_url)
        host = parsed.hostname
        encoded_remote_path = parsed.path or ""
        remote_path = unquote(encoded_remote_path)
        url_user = unquote(parsed.username or "")
        if not host or not remote_path:
            raise ValueError(f"无效 FTP 地址：{ftp_url}")
        filename = os.path.basename(remote_path)
        remote_dir = os.path.dirname(remote_path) or "/"
        encoded_filename = os.path.basename(encoded_remote_path)
        encoded_remote_dir = os.path.dirname(encoded_remote_path) or "/"
        if not filename.lower().endswith(".zip"):
            raise ValueError(f"FTP 地址不是 .zip 文件：{ftp_url}")
        return host, remote_dir, filename, remote_path, url_user, encoded_remote_dir, encoded_filename, encoded_remote_path

    def create_temp_folder(self, source_file_path):
        temp_folder = os.path.join(os.path.dirname(source_file_path), "temp")
        os.makedirs(temp_folder, exist_ok=True)
        return temp_folder

    def download_ftp_file_binary_threaded(self, ftp_url, temp_folder, file_index, total_files):
        (
            host,
            remote_dir,
            filename,
            remote_path,
            url_user,
            encoded_remote_dir,
            encoded_filename,
            encoded_remote_path,
        ) = self.parse_ftp_url(ftp_url)
        local_path = os.path.join(temp_folder, filename)
        login_user = url_user or FTP_USER

        self.thread_log("")
        self.thread_log(f"下载任务 {file_index}/{total_files}")
        self.thread_log(f"原始 FTP URL：{ftp_url}")
        self.thread_log(f"FTP 服务器 IP/Host：{host}")
        self.thread_log(f"FTP 远程目录（解码后，用于 cwd）：{remote_dir}")
        self.thread_log(f"FTP 远程目录（URL 原始编码）：{encoded_remote_dir}")
        self.thread_log(f"ZIP 文件名（解码后）：{filename}")
        self.thread_log(f"ZIP 文件名（URL 原始编码）：{encoded_filename}")
        self.thread_log(f"FTP 远程完整路径（解码后）：{remote_path}")
        self.thread_log(f"FTP 远程完整路径（URL 原始编码）：{encoded_remote_path}")
        self.thread_log(f"本地暂存路径：{local_path}")
        self.thread_log(f"FTP 登录用户：{login_user}")
        self.thread_log("FTP 密码：已配置，日志中不打印明文。")

        ftp = None
        try:
            self.thread_log("正在建立 FTP 连接，超时时间：120 秒...")
            ftp = FTP(host, timeout=120)
            self.thread_log("FTP 连接建立成功，正在登录...")
            ftp.login(login_user, FTP_PASS)
            self.thread_log("FTP 登录服务器成功。")

            try:
                self.thread_log(f"FTP 登录后的当前目录：{ftp.pwd()}")
            except Exception:
                self.thread_log("FTP 当前目录读取失败。")

            cwd_success = False
            cwd_errors = []
            # Normal FTP path should be decoded.  Keep encoded path only as a
            # fallback because some non-standard servers may expose literal %xx.
            for candidate_dir, label in [
                (remote_dir, "解码后目录"),
                (encoded_remote_dir, "URL原始编码目录"),
            ]:
                if not candidate_dir:
                    continue
                try:
                    self.thread_log(f"正在切换到 FTP 文件目录（{label}）：{candidate_dir}")
                    ftp.cwd(candidate_dir)
                    self.thread_log(f"FTP 目录切换成功，当前目录：{ftp.pwd()}")
                    cwd_success = True
                    break
                except Exception as e:
                    cwd_errors.append(f"{label}={candidate_dir} -> {e}")
                    self.thread_log(f"FTP 目录切换失败（{label}）：{e}")

            if not cwd_success:
                self.thread_log("所有 FTP 目录切换方式均失败。失败详情：")
                for err in cwd_errors:
                    self.thread_log(f"  {err}")
                self.thread_log("将尝试使用完整远程路径直接 RETR 下载，避免依赖 cwd。")

            active_retr_name = filename if cwd_success else remote_path
            active_size_name = filename if cwd_success else remote_path

            try:
                if cwd_success:
                    names = ftp.nlst()
                    self.thread_log(f"当前目录文件数量：{len(names)}")
                    if filename in names:
                        self.thread_log("目标 ZIP 文件在当前目录中确认存在。")
                    else:
                        self.thread_log("警告：目录列表中未直接看到目标 ZIP 文件，仍将尝试 RETR 下载。")
                else:
                    self.thread_log("未进入目标目录，跳过目录列表读取。")
            except Exception as e:
                self.thread_log(f"FTP 目录列表读取失败，将继续尝试下载：{e}")

            try:
                file_size = ftp.size(active_size_name)
            except Exception:
                file_size = 0

            self.thread_log(f"准备下载 ZIP 文件：{filename}")
            self.thread_log(f"FTP RETR 使用路径：{active_retr_name}")
            if file_size:
                self.thread_log(f"文件大小：{file_size} bytes")
            else:
                self.thread_log("无法获取文件大小，将显示任务级进度。")

            downloaded = 0
            with open(local_path, "wb") as f:
                def callback(data):
                    nonlocal downloaded
                    f.write(data)
                    downloaded += len(data)
                    if file_size > 0:
                        file_percent = downloaded / file_size * 100
                        total_percent = ((file_index - 1) + file_percent / 100) / total_files * 100
                        self.thread_progress(total_percent, f"正在下载 {filename}：{file_percent:.1f}%")
                ftp.retrbinary(f"RETR {active_retr_name}", callback, blocksize=1024 * 128)

            self.thread_log(f"下载完成：{local_path}")
            self.thread_log(f"已下载字节数：{downloaded} bytes")
            total_percent = file_index / total_files * 100
            self.thread_progress(total_percent, f"已完成 {file_index}/{total_files} 个文件")
            return local_path
        finally:
            if ftp is not None:
                try:
                    ftp.quit()
                    self.thread_log("FTP 连接已关闭。")
                except Exception:
                    try:
                        ftp.close()
                    except Exception:
                        pass
                    self.thread_log("FTP 连接已强制关闭。")

    def download_all_ftp_files_worker(self, source_file_path, device_code="", auto_upload=False):
        try:
            self.thread_log("开始后台下载 ZIP 文件...")
            self.thread_log(f"Source Excel 路径：{source_file_path}")
            self.thread_progress(0, "正在读取 Source Excel...")

            # First use normal field mapping; then scan the whole workbook as a fallback.
            # This guarantees that full FTP URLs in columns such as 整包地址 are found.
            try:
                rows, mapping = self.read_source_excel(source_file_path, use_gui_log=False)
                self.thread_log(f"Excel 字段映射识别结果：{self.format_mapping_for_log(mapping)}")
                mapped_urls = self.get_unique_ftp_urls(rows)
                self.thread_log(f"通过字段映射发现 FTP ZIP 文件数量：{len(mapped_urls)}")
            except Exception as e:
                mapped_urls = []
                self.thread_log(f"字段映射读取失败，将继续执行全表扫描：{e}")

            scanned_urls = self.scan_ftp_zip_urls_from_excel(source_file_path)
            unique_urls = list(dict.fromkeys(mapped_urls + scanned_urls))

            self.thread_log("")
            self.thread_log("========== FTP 下载任务 ==========")
            self.thread_log(f"需要下载的 ZIP 文件总数：{len(unique_urls)}")

            if not unique_urls:
                self.thread_progress(100, "没有发现需要下载的 FTP 文件")
                self.log_queue.put(("done", 0))
                return

            self.thread_log("待下载文件列表：")
            for index, url in enumerate(unique_urls, start=1):
                self.thread_log(f"  {index}. {url}")

            temp_folder = self.create_temp_folder(source_file_path)
            self.thread_log(f"临时下载目录：{temp_folder}")

            downloaded_files = []
            failed_files = []
            for index, ftp_url in enumerate(unique_urls, start=1):
                try:
                    local_path = self.download_ftp_file_binary_threaded(ftp_url, temp_folder, index, len(unique_urls))
                    downloaded_files.append(local_path)
                    self.thread_log(f"下载成功 [{index}/{len(unique_urls)}]：{os.path.basename(local_path)}")
                except Exception as e:
                    failed_files.append((ftp_url, str(e)))
                    self.thread_log(f"下载失败 [{index}/{len(unique_urls)}]：{ftp_url}")
                    self.thread_log(f"失败原因：{e}")
                    self.thread_log(traceback.format_exc())

            self.thread_progress(100, "全部 ZIP 下载任务完成")
            self.thread_log("")
            self.thread_log("========== 下载完成 ==========")
            self.thread_log(f"需要下载 ZIP 文件总数：{len(unique_urls)}")
            self.thread_log(f"成功下载 ZIP 文件数：{len(downloaded_files)}")
            self.thread_log(f"失败下载 ZIP 文件数：{len(failed_files)}")
            if failed_files:
                self.thread_log("失败文件列表：")
                for ftp_url, reason in failed_files:
                    self.thread_log(f"  {ftp_url}")
                    self.thread_log(f"    {reason}")

            if auto_upload and device_code and downloaded_files:
                temp_folder = self.create_temp_folder(source_file_path)
                self.thread_log("")
                self.thread_log("自动执行：ZIP 下载完成，开始 S3 上传...")
                self.upload_to_s3_worker(temp_folder, device_code)
            else:
                self.log_queue.put(("done", len(downloaded_files)))

        except Exception as e:
            self.thread_log("后台下载任务失败。")
            self.thread_log(str(e))
            self.thread_log(traceback.format_exc())
            self.log_queue.put(("error", str(e)))

    # ==========================================================
    # Data
    # ==========================================================

    def map_ota_type(self, value):
        """Normalize OTA upgrade type from the source Excel.

        Supported source values include:
        - 1 / 1.0 / 正常升级 / 1-正常升级 / 1 - 正常升级
        - 2 / 2.0 / 强制升级 / 2-强制升级 / 2 - 强制升级
        - 11 / 11.0 / FTE强制升级 / 11-FTE强制升级 / 11 - FTE强制升级

        The output template requires numeric codes only: 1, 2, or 11.
        """
        if value is None:
            return "1"

        raw = str(value).strip()
        if raw == "":
            return "1"

        # Normalize common punctuation and spacing variations from Excel cells.
        normalized = raw.replace("－", "-").replace("–", "-").replace("—", "-")
        normalized = normalized.replace("：", ":").replace("，", ",")
        normalized_no_space = re.sub(r"\s+", "", normalized)
        lowered = normalized_no_space.lower()

        # Numeric values may come from Excel as 1, 2, 11, 1.0, 2.0, 11.0.
        numeric_candidate = normalized_no_space
        if re.fullmatch(r"\d+(?:\.0+)?", numeric_candidate):
            numeric_code = str(int(float(numeric_candidate)))
            if numeric_code in ("1", "2", "11"):
                return numeric_code

        # Compound values like "11-FTE强制升级", "2 - 强制升级", "1:正常升级".
        prefix_match = re.match(r"^(11|2|1)(?:[-_:：、,，]|$)", normalized_no_space, re.IGNORECASE)
        if prefix_match:
            return prefix_match.group(1)

        # Text values. Check FTE before generic 强制, otherwise FTE强制升级 would map to 2.
        if "fte" in lowered and "强制" in normalized_no_space:
            return "11"
        if "强制" in normalized_no_space:
            return "2"
        if "正常" in normalized_no_space:
            return "1"

        # English fallbacks.
        if "force" in lowered or "mandatory" in lowered:
            if "fte" in lowered:
                return "11"
            return "2"
        if "normal" in lowered or "regular" in lowered:
            return "1"

        # Keep previous behavior: default to normal upgrade. Caller logs raw value.
        return "1"

    def extract_zip_filename(self, url_or_path):
        if not url_or_path:
            return ""
        text = str(url_or_path).strip()
        match = re.search(r"([^/\\]+\.zip)(?:[?#].*)?$", text, re.IGNORECASE)
        if match:
            return unquote(match.group(1)).strip()
        parsed = urlparse(text)
        filename = os.path.basename(unquote(parsed.path or text))
        return filename.strip() if filename.lower().endswith(".zip") else ""

    def find_zip_source_url_in_row(self, row):
        """Find the FTP/HTTP/local ZIP source value for one source Excel row.

        Normal mapping should read 升级包地址, but some files name the column 整包地址
        or other variants.  This fallback scans every original cell in that row so the
        generated 升级文件地址 can always include the exact ZIP filename.
        """
        candidates = [
            self.get_value(row, "升级包地址"),
            self.get_value(row, "升级文件地址"),
            self.get_value(row, "整包地址"),
        ]
        candidates.extend(row.get("_all_cell_values", []) or [])
        for value in candidates:
            text = str(value or "").strip()
            if text and ".zip" in text.lower():
                filename = self.extract_zip_filename(text)
                if filename:
                    return text
        return ""

    def normalize_path_part(self, path):
        if not path:
            return ""
        return re.sub(r"/+", "/", str(path).strip().replace("\\", "/"))

    def sanitize_url(self, url):
        if not url:
            return url
        url = re.sub(r'^(https?:)/+', r'\1//', url, flags=re.IGNORECASE)
        scheme_end = url.find("://")
        if scheme_end != -1:
            scheme = url[:scheme_end + 3]
            rest = url[scheme_end + 3:]
            rest = re.sub(r'/+', '/', rest)
            url = scheme + rest
        else:
            url = re.sub(r'/+', '/', url)
        return url

    def build_upgrade_url(self, source_upgrade_url, aws_suffix):
        source_upgrade_url = source_upgrade_url or ""
        zip_filename = self.extract_zip_filename(source_upgrade_url)
        clean_suffix = self.normalize_path_part(aws_suffix or self.get_default_aws_suffix())
        if not clean_suffix.startswith("/"):
            clean_suffix = "/" + clean_suffix
        if not clean_suffix.endswith("/"):
            clean_suffix += "/"
        final_url = AWS_PREFIX.rstrip("/") + clean_suffix + zip_filename
        final_url = self.sanitize_url(final_url)
        return final_url, zip_filename

    def build_target_version(self, source_version, target_version, version_change):
        if version_change is not None and str(version_change).strip() != "":
            return str(version_change).strip(), False
        source_text = "" if source_version is None else str(source_version).strip()
        target_text = "" if target_version is None else str(target_version).strip()
        source_suffix = source_text[-5:] if len(source_text) >= 5 else source_text
        target_suffix = target_text[-5:] if len(target_text) >= 5 else target_text
        return f"{source_suffix}_to_{target_suffix}", True

    def safe_filename_part(self, text):
        text = str(text or "").strip()
        text = re.sub(r'[\\/:*?"<>|]+', "_", text)
        return text or "Device"

    def generate_output_path(self, source_file_path, device_code, row_count):
        folder = os.path.dirname(source_file_path)
        device_part = self.safe_filename_part(device_code)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(folder, f"{device_part}_{row_count}_{timestamp}.xlsx")

    def validate_output_row(self, row_index, row_data):
        warnings = []
        upgrade_url = row_data.get("升级文件地址", "")
        if upgrade_url:
            if not re.match(r'^https?://', upgrade_url, re.IGNORECASE):
                warnings.append(f"  ⚠ 升级文件地址 不是合法 URL（缺少 http/https 协议）：{upgrade_url}")
            if re.search(r'/{3,}', upgrade_url):
                warnings.append(f"  ⚠ 升级文件地址 包含非法连续斜杠（///）：{upgrade_url}")
            if not upgrade_url.lower().endswith(".zip"):
                warnings.append(f"  ⚠ 升级文件地址 不以 .zip 结尾：{upgrade_url}")
        else:
            warnings.append("  ⚠ 升级文件地址 为空")

        ota_type = str(row_data.get("OTA类型", "")).strip()
        if ota_type not in ("1", "2", "11"):
            warnings.append(f"  ⚠ OTA类型 值无效（期望 1 / 2 / 11，实际：'{ota_type}'）")

        if warnings:
            self.log_message(f"  [校验] 第 {row_index} 行发现问题：")
            for w in warnings:
                self.log_message(w)
        else:
            self.log_message(f"  [校验] 第 {row_index} 行校验通过 ✓")
        return warnings

    def get_value(self, row, field):
        return row.get(field, "")

    def format_mapping_for_log(self, mapping):
        if not mapping:
            return "未识别到任何字段"
        return ", ".join([f"{field}->{get_column_letter(col_idx)}" for field, col_idx in sorted(mapping.items(), key=lambda x: x[1])])

    def log_required_field_status(self, mapping):
        required_fields = ["源版本", "目标版本", "OTA类型", "升级包地址", "升级文件的MD5值", "文件大小"]
        self.log_message("关键字段检查：")
        for field in required_fields:
            if field in mapping:
                self.log_message(f"  ✓ 已识别：{field} -> {get_column_letter(mapping[field])}列")
            else:
                self.log_message(f"  ⚠ 未识别：{field}。生成时该字段将留空或使用默认规则。")
        if "区域" in mapping:
            self.log_message(f"  ✓ 已识别：区域 -> {get_column_letter(mapping['区域'])}列，将复制到输出文件的“区域组”。")
        else:
            self.log_message("  ⚠ 未识别：区域。输出文件“区域组”将留空。")
        if "MAC" in mapping:
            self.log_message(f"  说明：已检测到 Source Excel 的 MAC 列 {get_column_letter(mapping['MAC'])}，但按照规则不会复制；输出 MAC组只使用 UI 输入。")

    # ==========================================================
    # Workbook
    # ==========================================================

    def clean_file_size(self, value):
        if not value:
            return value
        return re.sub(r"[,\s_]", "", str(value).strip())

    def validate_feature_code(self, value):
        if not value:
            return value
        digits_only = re.sub(r"\D", "", str(value))
        if not digits_only:
            return value
        if len(digits_only) < 24:
            digits_only = digits_only.zfill(24)
        elif len(digits_only) > 24:
            digits_only = digits_only[-24:]
        return digits_only

    def apply_text_format(self, cell):
        cell.number_format = "@"
        cell.alignment = Alignment(vertical="center")

    def create_workbook(self, output_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "升级规则"

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.number_format = "@"

        for row_idx, row_data in enumerate(output_rows, start=2):
            for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                self.apply_text_format(cell)
                cell.border = border

        ws.freeze_panes = "A2"

        # ===============================
        # OTA升级提示语 Sheet
        # ===============================
        msg_ws = wb.create_sheet("升级提示语")

        msg_headers = ["语种", "描述"]
        for col_idx, header in enumerate(msg_headers, start=1):
            cell = msg_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ota_messages = [
            ["简体中文", "有新软件版本可用。更新包括：\n-改善用户体验\n-其他一些小问题\n\n请更新电视软件以改善用户体验。\n软件升级时，请勿断电。"],
            ["法语", "Une nouvelle version du logiciel est disponible. Les mises à jour incluent :\n-Améliorer l’expérience utilisateur\n- Quelques corrections mineures\n\nVeuillez mettre à jour le logiciel du téléviseur pour améliorer votre expérience utilisateur.\nNe coupez pas l’alimentation lors de la mise à niveau du logiciel."],
            ["西班牙语", "Una nueva versión de software está disponible. Las actualizaciones incluyen:\n-Mejorar la experiencia del usuario\n-Algunas otras correcciones menores\n\nActualice el software de TV para mejorar su experiencia de usuario.\nCuando el S/W se está actualizando, NO apague la alimentación."],
            ["繁体中文", "有新軟體版本可用。更新包括：\n-改善用戶體驗\n-其他一些小問題\n\n請更新電視軟體以改善用戶體驗。\n軟體升級時，請勿斷電。"],
            ["德语", "Eine neue Softwareversion ist verfügbar. Aktualisierungen umfassen:\n- Verbesserung des Nutzungserlebnisses\n- weitere Bugfixes\n\nBitte aktualisieren Sie die TV-Software, um Ihre Benutzererfahrung zu verbessern.\nSchalten Sie die Stromversorgung NICHT aus, wenn die Software aktualisiert wird."],
            ["葡萄牙语", "Está disponível uma atualização de software importante.\nO que há de novo?\n-Correções de bugs de software\n-Melhorias de plataforma\n\nA alimentação tem de estar sempre ligada até à conclusão da atualização."]
        ]

        for row_idx, row in enumerate(ota_messages, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = msg_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        msg_ws.column_dimensions["A"].width = 20
        msg_ws.column_dimensions["B"].width = 120


        for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
            max_len = len(header)
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 90)
        return wb

    # ==========================================================
    # Validation
    # ==========================================================

    def validate_inputs(self, source_file_path):
        if not source_file_path:
            raise ValueError("请先选择 Source Excel 文件。")
        if not os.path.exists(source_file_path):
            raise ValueError("Source Excel 文件不存在。")
        if not source_file_path.lower().endswith((".xlsx", ".xlsm")):
            raise ValueError("请选择 .xlsx 或 .xlsm 格式的 Excel 文件。")

    def prepare_common(self, animation_text):
        self.log_message("")
        self.log_message("=" * 60)
        self.log_message(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {animation_text}")
        self.log_message("=" * 60)

        self.download_progress_var.set(0)
        self.download_status_var.set("下载进度：0%")

        source_file_path = self.source_file_var.get().strip()
        self.log_message(f"Source Excel 路径：{source_file_path}")
        self.validate_inputs(source_file_path)
        self.log_message("输入文件校验通过。")
        self.update_aws_suffix_from_device_code()
        self.log_message(f"设备类型代码：{self.device_code_var.get().strip() or '(空)'}")
        self.log_message(f"设备类型名称：{self.device_name_var.get().strip() or '(空)'}")
        self.log_message(f"Feature Code / 特征码：{self.feature_code_var.get().strip() or '(空)'}")
        self.log_message(f"MAC 地址组：{self.mac_group_var.get().strip() or '(空)'}")
        self.log_message(f"AWS 下载地址前缀：{AWS_PREFIX}")
        self.log_message(f"AWS 文件路径后缀：{self.aws_suffix_var.get().strip()}")
        self.log_message(f"AWS UI 完整显示路径：{self.aws_full_location_var.get().strip()}")
        self.save_config()
        self.log_message(f"配置已保存：{CONFIG_FILE}")
        self.disable_all_buttons()
        self.start_animation(animation_text)
        return source_file_path

    # ==========================================================
    # Buttons
    # ==========================================================

    def generate_template_only(self):
        self.generate_excel_template()

    def download_zip_only(self):
        try:
            source_file_path = self.prepare_common("正在下载ZIP文件")
            device_code_for_zip = self.device_code_var.get().strip()
            self.log_message("按钮：下载ZIP文件")
            self.log_message("ZIP 下载任务已启动，正在后台运行。本操作只下载到本地 temp 目录，不会上传 S3。")
            self.log_message("界面不会卡住，可以继续查看日志和进度。")
            worker = threading.Thread(
                target=self.download_all_ftp_files_worker,
                args=(source_file_path, device_code_for_zip, False),
                daemon=True
            )
            worker.start()
        except Exception as e:
            self.stop_animation("下载失败")
            self.log_message(str(e))
            self.log_message(traceback.format_exc())
            self.show_center_error("错误", str(e))

    def auto_execute_all(self):
        self.generate_excel_template(start_zip_after=True)

    def upload_to_s3_only(self):
        try:
            source_file_path = self.prepare_common("正在上传至S3")
            device_code = self.device_code_var.get().strip()
            if not device_code:
                raise ValueError("请先填写 设备类型代码，S3 路径需要该字段。")
            temp_folder = os.path.join(os.path.dirname(source_file_path), "temp")
            self.log_message("按钮：上传ZIP文件到AWS")
            self.log_message("S3 上传任务已启动，正在后台运行。")
            self.log_message("界面不会卡住，可以继续查看日志和进度。")
            worker = threading.Thread(target=self.upload_to_s3_worker, args=(temp_folder, device_code), daemon=True)
            worker.start()
        except Exception as e:
            self.stop_animation("上传失败")
            self.log_message(str(e))
            self.log_message(traceback.format_exc())
            self.show_center_error("错误", str(e))

    def check_aws_cli_ready(self):
        try:
            aws_path = shutil.which("aws")
            if aws_path is None:
                self.thread_log("AWS CLI is not installed or not found in PATH.")
                self.thread_log("请先安装 AWS CLI，并确认 Terminal 中可以执行：aws --version")
                return False

            self.thread_log(f"AWS CLI 路径：{aws_path}")
            version_result = subprocess.run(["aws", "--version"], capture_output=True, text=True, timeout=20)
            version_text = (version_result.stdout or version_result.stderr).strip()
            if version_text:
                self.thread_log(f"AWS CLI 版本：{version_text}")
            result = subprocess.run(["aws", "sts", "get-caller-identity"], capture_output=True, text=True, timeout=20)
            if result.returncode == 0:
                self.thread_log("AWS CLI is configured and ready.")
                identity = result.stdout.strip()
                if identity:
                    self.thread_log(f"AWS 当前身份：{identity}")
                return True

            self.thread_log("AWS CLI is not configured correctly. Please run `aws configure` first.")
            if result.stderr.strip():
                self.thread_log(f"AWS 错误信息：{result.stderr.strip()}")
            elif result.stdout.strip():
                self.thread_log(f"AWS 输出信息：{result.stdout.strip()}")
            return False

        except subprocess.TimeoutExpired:
            self.thread_log("AWS CLI 检查超时，请检查网络、AWS CLI 配置或权限。")
            return False
        except Exception as e:
            self.thread_log("AWS CLI 检查异常，无法继续执行 S3 操作。")
            self.thread_log(str(e))
            return False

    # ==========================================================
    # S3 Upload
    # ==========================================================

    def build_s3_destination(self, device_code):
        """Return the S3 destination prefix for today's upload.
        Format: s3://fam-media-andr/ota/{device_code}/{YYYYMMDD}/
        """
        today = datetime.now().strftime("%Y%m%d")
        device_safe = device_code.strip() or "unknown"
        return f"s3://fam-media-andr/ota/{device_safe}/{today}/"

    def upload_to_s3_worker(self, temp_folder, device_code):
        try:
            self.thread_log("")
            self.thread_log("========== S3 上传任务 ==========")
            self.thread_log("正在检查 AWS CLI 安装和配置状态...")

            if not self.check_aws_cli_ready():
                self.thread_log("S3 上传已停止：AWS CLI 未准备好。")
                self.log_queue.put(("s3_done", 0, 0))
                return

            if not os.path.isdir(temp_folder):
                self.thread_log(f"临时目录不存在，请先执行 下载ZIP文件：{temp_folder}")
                self.log_queue.put(("s3_done", 0, 0))
                return

            zip_files = [os.path.join(temp_folder, f) for f in os.listdir(temp_folder) if f.lower().endswith(".zip")]
            if not zip_files:
                self.thread_log(f"临时目录中未找到任何 .zip 文件：{temp_folder}")
                self.log_queue.put(("s3_done", 0, 0))
                return

            s3_dest = self.build_s3_destination(device_code)
            self.thread_log(f"找到 {len(zip_files)} 个 ZIP 文件")
            self.thread_log(f"S3 目标路径：{s3_dest}")
            self.thread_log("")

            success_count = 0
            fail_count = 0
            total = len(zip_files)

            for idx, local_path in enumerate(zip_files, start=1):
                filename = os.path.basename(local_path)
                s3_target = s3_dest + filename
                self.thread_log(f"[{idx}/{total}] 正在上传：{filename}")
                self.thread_log(f"  本地：{local_path}")
                self.thread_log(f"  目标：{s3_target}")
                self.thread_progress((idx - 1) / total * 100, f"正在上传 {filename}（{idx}/{total}）")

                try:
                    cmd = ["aws", "s3", "cp", local_path, s3_target, "--no-progress"]
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
                    if result.returncode == 0:
                        self.thread_log(f"  ✓ 上传成功：{filename}")
                        if result.stdout.strip():
                            self.thread_log(f"  {result.stdout.strip()}")
                        success_count += 1
                    else:
                        self.thread_log(f"  ✗ 上传失败：{filename}")
                        if result.stderr.strip():
                            self.thread_log(f"  错误信息：{result.stderr.strip()}")
                        fail_count += 1
                except subprocess.TimeoutExpired:
                    self.thread_log(f"  ✗ 上传超时（10分钟）：{filename}")
                    fail_count += 1
                except FileNotFoundError:
                    self.thread_log("  ✗ 未找到 aws 命令，请确认已安装 AWS CLI 并配置好 PATH。")
                    self.thread_log("  安装参考：https://docs.aws.amazon.com/cli/latest/userguide/install-cliv2.html")
                    fail_count += 1
                    break
                except Exception as e:
                    self.thread_log(f"  ✗ 上传异常：{e}")
                    fail_count += 1

            self.thread_progress(100, f"S3 上传完成：成功 {success_count} / 失败 {fail_count}")
            self.thread_log("")
            self.thread_log("========== S3 上传结果 ==========")
            self.thread_log(f"S3 目标路径：{s3_dest}")
            self.thread_log(f"成功上传：{success_count} 个文件")
            if fail_count:
                self.thread_log(f"上传失败：{fail_count} 个文件")
            self.log_queue.put(("s3_done", success_count, fail_count))

        except Exception as e:
            self.thread_log("S3 上传任务异常终止。")
            self.thread_log(str(e))
            self.thread_log(traceback.format_exc())
            self.log_queue.put(("error", str(e)))

    # ==========================================================
    # Generate Excel
    # ==========================================================

    def generate_excel_template(self, start_zip_after=False):
        try:
            source_file_path = self.prepare_common("正在自动执行" if start_zip_after else "正在生成模版文件")
            device_code = self.device_code_var.get().strip()
            device_name = self.device_name_var.get().strip()
            feature_code = self.feature_code_var.get().strip()
            aws_suffix = self.get_default_aws_suffix(device_code)
            self.aws_suffix_var.set(aws_suffix)
            self.aws_full_location_var.set(AWS_PREFIX.rstrip("/") + aws_suffix)
            mac_group = self.mac_group_var.get().strip()

            self.log_message("按钮：自动执行" if start_zip_after else "按钮：生成模版文件")
            self.log_message("开始生成 OTA 部署文件...")
            self.log_message(f"输出文件将保存到 Source Excel 同级目录：{os.path.dirname(source_file_path)}")

            rows, mapping = self.read_source_excel(source_file_path, use_gui_log=True)
            self.log_message(f"是否检测到“版本变化”列：{'是' if '版本变化' in mapping else '否'}")
            self.log_message(f"是否检测到“区域”列：{'是，将复制到输出“区域组”' if '区域' in mapping else '否，输出“区域组”将留空'}")
            self.log_message(f"是否检测到 Source Excel 的 MAC 列：{'是，但不会复制' if 'MAC' in mapping else '否'}")
            output_rows = []
            auto_target_version_count = 0

            for index, row in enumerate(rows, start=1):
                source_version = self.get_value(row, "源版本")
                full_target_version = self.get_value(row, "目标版本")
                version_change = self.get_value(row, "版本变化")
                ota_type_raw = self.get_value(row, "OTA类型")
                mapped_upgrade_url = self.get_value(row, "升级包地址")
                source_upgrade_url = self.find_zip_source_url_in_row(row)
                region_group = self.get_value(row, "区域")

                target_version, auto_generated = self.build_target_version(source_version, full_target_version, version_change)
                if auto_generated:
                    auto_target_version_count += 1

                ota_type = self.map_ota_type(ota_type_raw)
                upgrade_url, zip_filename = self.build_upgrade_url(source_upgrade_url, aws_suffix)
                if not zip_filename:
                    self.log_message(f"  ⚠ 第 {index} 行未能识别 ZIP 文件名；请检查源 Excel 是否包含 整包地址/升级包地址 且值以 .zip 结尾。")

                output_rows.append({
                    "设备类型代码": device_code,
                    "设备类型名称": device_name,
                    "特征码": self.validate_feature_code(feature_code),
                    "内部机型信息": self.get_value(row, "机型信息"),
                    "设备扩展信息": self.get_value(row, "机器扩展码"),
                    "OTA类型": ota_type,
                    "源版本": source_version,
                    "目标完整版本": full_target_version,
                    "目标版本": target_version,
                    "品牌组": self.get_value(row, "品牌"),
                    "区域组": region_group,
                    "MAC组": mac_group,
                    "定向组": "",
                    "升级文件地址": upgrade_url,
                    "升级文件的MD5值": self.get_value(row, "升级文件的MD5值"),
                    "文件大小": self.clean_file_size(self.get_value(row, "文件大小")),
                    "SHA256": self.get_value(row, "SHA256"),
                    "EULA文件地址": self.get_value(row, "EULA文件地址"),
                })

                self.log_message("")
                self.log_message(f"第 {index} 行")
                self.log_message(f"  源 Excel 行号：{row.get('_row_number', '')}")
                self.log_message(f"  内部机型信息：{self.get_value(row, '机型信息')}")
                self.log_message(f"  设备扩展信息：{self.get_value(row, '机器扩展码')}")
                self.log_message(f"  OTA类型：{ota_type_raw} -> {ota_type}")
                self.log_message(f"  源版本：{source_version}")
                self.log_message(f"  目标完整版本：{full_target_version}")
                self.log_message(f"  目标版本：{target_version} {'(自动生成)' if auto_generated else '(来自版本变化列)'}")
                self.log_message(f"  品牌组：{self.get_value(row, '品牌')}")
                self.log_message(f"  区域组：{region_group}")
                self.log_message(f"  MAC组：{mac_group}（只使用 UI 输入，不复制 Source Excel MAC 列）")
                self.log_message(f"  Source 映射升级包地址：{mapped_upgrade_url}")
                self.log_message(f"  Source 实际 ZIP 地址/路径：{source_upgrade_url}")
                self.log_message(f"  ZIP 文件：{zip_filename}")
                self.log_message(f"  升级文件地址：{upgrade_url}")
                self.log_message(f"  MD5：{self.get_value(row, '升级文件的MD5值')}")
                self.log_message(f"  文件大小：{self.clean_file_size(self.get_value(row, '文件大小'))}")
                self.log_message(f"  SHA256：{self.get_value(row, 'SHA256')}")
                self.validate_output_row(index, output_rows[-1])

            output_path = self.generate_output_path(source_file_path, device_code, len(output_rows))
            wb = self.create_workbook(output_rows)
            wb.save(output_path)

            self.log_message("")
            self.log_message("========== 生成结果 ==========")
            self.log_message(f"生成表格行数：{len(output_rows)}")
            self.log_message(f"自动生成目标版本行数：{auto_target_version_count}")
            self.log_message(f"区域组复制状态：已从 Source Excel 的“区域”字段复制到输出“区域组”（如源文件未识别该列则为空）。")
            self.log_message(f"输出文件路径：{output_path}")
            self.log_message("模版文件生成成功。")

            if start_zip_after:
                self.status_var.set("模版完成，开始后台下载ZIP...")
                self.log_message("")
                self.log_message("自动执行：开始后台下载 ZIP 文件。")
                worker = threading.Thread(
                    target=self.download_all_ftp_files_worker,
                    args=(source_file_path, device_code, True),
                    daemon=True
                )
                worker.start()
            else:
                self.stop_animation("模版生成成功")
                self.show_center_message("生成完成", f"Excel 文件生成成功。\n\n输出文件：\n{output_path}")

        except Exception as e:
            self.stop_animation("生成失败")
            self.log_message("")
            self.log_message("生成失败。")
            self.log_message(str(e))
            self.log_message(traceback.format_exc())
            self.show_center_error("错误", str(e))


# ==========================================================
# Main
# ==========================================================

def main():
    root = tk.Tk()
    OTATemplateApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
